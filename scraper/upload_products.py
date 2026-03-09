"""
One-time script to import products from 'Final Comp data.xlsx' into pm_products.

Usage:
  export SUPABASE_URL=https://xxx.supabase.co
  export SUPABASE_SERVICE_KEY=eyJ...
  python upload_products.py /path/to/Final\ Comp\ data.xlsx
"""

from __future__ import annotations

import re
import sys
import logging

import os

import httpx
import openpyxl

SUPABASE_URL = os.environ.get("SUPABASE_URL", "")
SUPABASE_SERVICE_KEY = os.environ.get("SUPABASE_SERVICE_KEY", "")
DB_BATCH_SIZE = 500

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(message)s")
logger = logging.getLogger(__name__)


def extract_amazon_asin(url: str | None) -> str | None:
    """Extract ASIN from Amazon URL like amazon.sa/dp/B08XYZ123"""
    if not url:
        return None
    m = re.search(r"/dp/([A-Z0-9]{10})", url)
    return m.group(1) if m else None


def extract_ninja_product_id(url: str | None) -> str | None:
    """Extract product slug from Ninja URL."""
    if not url:
        return None
    # URL like: ananinja.com/sa/en/product/some-product-slug
    m = re.search(r"/product/([^/?#]+)", url)
    return m.group(1) if m else None


def extract_lulu_product_id(url: str | None) -> str | None:
    """Extract product ID from Lulu URL like gcc.luluhypermarket.com/.../p/123456"""
    if not url:
        return None
    m = re.search(r"/p/(\d+)", url)
    return m.group(1) if m else None


_NA_VALUES = {"na", "n/a", "#n/a", "#na", "none", "null", "-", ""}


def clean_str(val) -> str | None:
    """Clean cell value to string or None. Treats NA-like values as None."""
    if val is None:
        return None
    s = str(val).strip()
    if s.lower() in _NA_VALUES:
        return None
    return s if s else None


def parse_excel(filepath: str) -> dict[str, dict]:
    """
    Parse all 3 sheets and merge into a single product dict keyed by SKU.
    Each sheet may have the same SKU with different competitor data.
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    products: dict[str, dict] = {}

    sheet_configs = [
        {
            "name": "Amazon Now",
            "competitor": "amazon",
            "link_col": "correct amazon now mapping link",
            "comp_sku_col": "correct amazon now sku",
            "title_col": "title",
            "brand_col": "brand",
            "top_col": "top 2500",
        },
        {
            "name": "Ninja",
            "competitor": "ninja",
            "link_col": "correct ninja mapping link",
            "comp_sku_col": "correct ninja sku id",
            "title_col": "product_title",
            "brand_col": "brand",
            "top_col": "top 2500 sku",
        },
        {
            "name": "Lulu",
            "competitor": "lulu",
            "link_col": "correct lulu mapping link",
            "comp_sku_col": "correct lulu sku id",
            "title_col": "product_title",
            "brand_col": "brand_code",
            "top_col": "top 2500",
        },
    ]

    for cfg in sheet_configs:
        sheet_name = cfg["name"]
        if sheet_name not in wb.sheetnames:
            logger.warning(f"Sheet '{sheet_name}' not found, skipping")
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Build header map (case-insensitive)
        raw_headers = [str(h).strip().lower() if h else "" for h in rows[0]]
        header_map = {h: i for i, h in enumerate(raw_headers) if h}

        def get_cell(row, col_name):
            idx = header_map.get(col_name.lower())
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        count = 0
        for row in rows[1:]:
            sku = clean_str(get_cell(row, "sku"))
            if not sku:
                continue

            # Initialize product if new SKU
            if sku not in products:
                products[sku] = {
                    "sku": sku,
                    "title": clean_str(get_cell(row, cfg["title_col"])) or "",
                    "category": clean_str(get_cell(row, "category")),
                    "brand": clean_str(get_cell(row, cfg["brand_col"])),
                    "is_top_2500": str(get_cell(row, cfg["top_col"])).strip().lower() in ("yes", "true", "1"),
                    "noon_debugger_url": clean_str(get_cell(row, "debuger link")),
                }

            # Add competitor-specific data
            comp = cfg["competitor"]
            comp_url = clean_str(get_cell(row, cfg["link_col"]))
            comp_sku = clean_str(get_cell(row, cfg["comp_sku_col"]))

            products[sku][f"{comp}_url"] = comp_url
            products[sku][f"{comp}_sku"] = comp_sku

            if comp == "amazon":
                products[sku]["amazon_asin"] = extract_amazon_asin(comp_url)
            elif comp == "ninja":
                products[sku]["ninja_product_id"] = extract_ninja_product_id(comp_url)
            elif comp == "lulu":
                products[sku]["lulu_product_id"] = extract_lulu_product_id(comp_url)

            count += 1

        logger.info(f"Parsed {count} rows from sheet '{sheet_name}'")

    wb.close()
    return products


ALL_KEYS = [
    "sku", "title", "category", "brand", "is_top_2500", "noon_debugger_url",
    "amazon_sku", "amazon_url", "amazon_asin",
    "ninja_sku", "ninja_url", "ninja_product_id",
    "lulu_sku", "lulu_url", "lulu_product_id",
]


def upload_to_supabase(products: dict[str, dict]) -> None:
    """Upload all products to pm_products table via REST API."""
    # Normalize: every row must have the same keys for PostgREST
    rows = []
    for p in products.values():
        row = {k: p.get(k) for k in ALL_KEYS}
        rows.append(row)
    logger.info(f"Uploading {len(rows)} unique products to Supabase...")

    headers = {
        "apikey": SUPABASE_SERVICE_KEY,
        "Authorization": f"Bearer {SUPABASE_SERVICE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "resolution=merge-duplicates",
    }
    url = f"{SUPABASE_URL}/rest/v1/pm_products?on_conflict=sku"

    with httpx.Client(timeout=60) as client:
        for i in range(0, len(rows), DB_BATCH_SIZE):
            batch = rows[i : i + DB_BATCH_SIZE]
            resp = client.post(url, json=batch, headers=headers)
            if resp.status_code not in (200, 201):
                logger.error(f"Batch {i} failed: {resp.status_code} {resp.text[:200]}")
            else:
                logger.info(f"Uploaded {min(i + DB_BATCH_SIZE, len(rows))}/{len(rows)}")

    logger.info("Upload complete!")


def main():
    if len(sys.argv) < 2:
        print("Usage: python upload_products.py <path_to_excel>")
        sys.exit(1)

    filepath = sys.argv[1]
    logger.info(f"Parsing {filepath}...")
    products = parse_excel(filepath)
    logger.info(f"Found {len(products)} unique SKUs across all sheets")

    upload_to_supabase(products)


if __name__ == "__main__":
    main()
